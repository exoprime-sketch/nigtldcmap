"""Parse the standardized Vietnam V124 Excel workbooks."""

from __future__ import annotations

import io
import re
from typing import Any, Iterable, Mapping

import openpyxl

from .normalization import (
    ELEMENT_ID_RE,
    credential_finding,
    decode_zip_filename,
    duplicate_summary,
    extract_element_id,
    is_explicit_placeholder_row,
    is_placeholder,
    nfc_text,
    normalize_field_name,
    normalize_row,
    normalize_value,
    semantic_attribute_name,
    stable_sort_key,
)


OBSERVATION_SHEET = "1.1_observation(측정값)"
ENTITY_SHEET = "1.2_entity(레코드형)"
METADATA_SHEET = "2_meta_info"
FRAMEWORK_SHEET = "db_framework"
CORE_SHEETS = (OBSERVATION_SHEET, ENTITY_SHEET, METADATA_SHEET, FRAMEWORK_SHEET)

MAX_RELEVANT_COLUMNS = 128
MAX_RELEVANT_ROWS = 250_000

_HEADER_ALIASES = {
    "요소_id": "element_id",
    "요소id": "element_id",
    "element_id": "element_id",
    "요소_명": "element_name",
    "요소명": "element_name",
    "indicator_id": "indicator_id",
    "country_iso3": "country_iso3",
    "year": "year",
    "value": "value",
    "missing_reason_code": "missing_reason_code",
    "note": "note",
    "lat": "lat",
    "latitude": "lat",
    "lon": "lon",
    "lng": "lon",
    "longitude": "lon",
    "geometry_type": "geometry_type",
    "crs": "crs",
    "요소_kr": "element_kr",
    "요소_en": "element_en",
}

_YEAR_RE = re.compile(r"(?<!\d)((?:19|20)\d{2})(?!\d)")


def _sheet_rows(sheet: Any) -> tuple[list[list[Any]], list[str]]:
    warnings: list[str] = []
    max_row = min(int(sheet.max_row or 0), MAX_RELEVANT_ROWS)
    max_col = min(int(sheet.max_column or 0), MAX_RELEVANT_COLUMNS)
    if int(sheet.max_row or 0) > MAX_RELEVANT_ROWS:
        warnings.append(f"row-limit-applied:{sheet.max_row}")
    if int(sheet.max_column or 0) > MAX_RELEVANT_COLUMNS:
        warnings.append(f"column-limit-applied:{sheet.max_column}")
    rows: list[list[Any]] = []
    for cells in sheet.iter_rows(
        min_row=1,
        max_row=max_row,
        min_col=1,
        max_col=max_col,
        values_only=False,
    ):
        values: list[Any] = []
        for cell in cells:
            if getattr(cell, "data_type", None) == "e":
                warnings.append(f"excel-error-cell:{cell.coordinate}:{cell.value}")
                values.append(None)
            else:
                values.append(cell.value)
        rows.append(normalize_row(values))
    return rows, warnings


def _canonical_header(value: Any) -> str:
    normalized = normalize_field_name(value)
    return _HEADER_ALIASES.get(normalized, normalized)


def _find_header(rows: list[list[Any]], required: set[str]) -> tuple[int | None, list[str]]:
    for index, row in enumerate(rows[:25]):
        headers = [_canonical_header(value) for value in row]
        if required.issubset(set(headers)):
            return index, headers
    return None, []


def _record_from_row(headers: list[str], row: list[Any]) -> dict[str, Any]:
    record: dict[str, Any] = {}
    for index, header in enumerate(headers):
        if not header:
            continue
        value = normalize_value(row[index]) if index < len(row) else None
        # Headers are unique in the current source. Preserve future duplicate
        # columns deterministically rather than silently overwriting them.
        key = header
        suffix = 2
        while key in record:
            key = f"{header}_{suffix}"
            suffix += 1
        record[key] = value
    return record


def _valid_source_record(record: Mapping[str, Any]) -> bool:
    element_id = extract_element_id(record.get("element_id"))
    return bool(element_id and record.get("indicator_id"))


def _normalized_element_id(record: dict[str, Any]) -> None:
    element_id = extract_element_id(record.get("element_id"))
    if element_id:
        record["element_id"] = element_id
    if record.get("country_iso3"):
        record["country_iso3"] = nfc_text(str(record["country_iso3"])).upper()


def _public_observation_populated(record: Mapping[str, Any]) -> bool:
    return not is_placeholder(record.get("value"))


def _public_entity_populated(record: Mapping[str, Any]) -> bool:
    coordinate_fields = ("lat", "lon", "geometry_type", "crs")
    if any(not is_placeholder(record.get(key)) for key in coordinate_fields):
        return True
    attrs = record.get("attributes") or {}
    return isinstance(attrs, Mapping) and any(not is_placeholder(value) for value in attrs.values())


def _security_scan(
    record: dict[str, Any],
    *,
    field_labels: Mapping[str, str] | None = None,
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    labels = field_labels or {}
    findings: list[dict[str, Any]] = []
    result: dict[str, Any] = {}
    for key, value in record.items():
        if isinstance(value, Mapping):
            nested, nested_findings = _security_scan(dict(value))
            result[key] = nested
            findings.extend({**item, "field": f"{key}.{item['field']}"} for item in nested_findings)
            continue
        finding = credential_finding(labels.get(key, key), value)
        if finding:
            findings.append(finding)
            result[key] = None
        else:
            result[key] = value
    return result, findings


def _parse_observations(rows: list[list[Any]]) -> dict[str, Any]:
    header_index, headers = _find_header(rows, {"element_id", "indicator_id", "value"})
    if header_index is None:
        return {
            "records": [],
            "populated": [],
            "missing": [],
            "supplementalRows": [],
            "placeholderRows": 0,
            "unclassifiedRows": 0,
            "templateRows": sum(bool(row) for row in rows),
            "errors": ["observation-header-not-found"],
            "securityFindings": [],
        }
    records: list[dict[str, Any]] = []
    placeholders = 0
    unclassified = 0
    supplemental_rows: list[dict[str, Any]] = []
    findings: list[dict[str, Any]] = []
    for row_index, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        if not row:
            continue
        candidate = _record_from_row(headers, row)
        if not _valid_source_record(candidate):
            if is_explicit_placeholder_row(row):
                placeholders += 1
            else:
                supplemental_rows.append({"source_row": row_index, "cells": row})
            continue
        _normalized_element_id(candidate)
        candidate["source_row"] = row_index
        candidate, row_findings = _security_scan(candidate)
        findings.extend({**item, "sourceRow": row_index} for item in row_findings)
        records.append(candidate)
    records.sort(key=stable_sort_key)
    populated = [record for record in records if _public_observation_populated(record)]
    missing = [record for record in records if not _public_observation_populated(record)]
    return {
        "records": records,
        "populated": populated,
        "missing": missing,
        "supplementalRows": supplemental_rows,
        "placeholderRows": placeholders,
        "unclassifiedRows": unclassified,
        "templateRows": sum(bool(row) for row in rows[: header_index + 1]),
        "errors": [],
        "securityFindings": findings,
    }


def _parse_entities(rows: list[list[Any]]) -> dict[str, Any]:
    header_index, headers = _find_header(rows, {"element_id", "indicator_id"})
    if header_index is None:
        return {
            "records": [],
            "populated": [],
            "missing": [],
            "supplementalRows": [],
            "placeholderRows": 0,
            "unclassifiedRows": 0,
            "templateRows": sum(bool(row) for row in rows),
            "attributeLabels": {},
            "errors": ["entity-header-not-found"],
            "securityFindings": [],
        }
    label_row: list[Any] = []
    data_start = header_index + 1
    if data_start < len(rows):
        possible_labels = rows[data_start]
        first = possible_labels[0] if possible_labels else None
        if not extract_element_id(first) and any(possible_labels):
            label_row = possible_labels
            data_start += 1
    attribute_labels: dict[str, str] = {}
    for index, header in enumerate(headers):
        if not header.startswith("attr_"):
            continue
        label = label_row[index] if index < len(label_row) else None
        attribute_labels[header] = nfc_text(str(label or header))
    records: list[dict[str, Any]] = []
    placeholders = 0
    unclassified = 0
    supplemental_rows: list[dict[str, Any]] = []
    findings: list[dict[str, Any]] = []
    for row_index, row in enumerate(rows[data_start:], start=data_start + 1):
        if not row:
            continue
        raw = _record_from_row(headers, row)
        if not _valid_source_record(raw):
            if is_explicit_placeholder_row(row):
                placeholders += 1
            else:
                supplemental_rows.append({"source_row": row_index, "cells": row})
            continue
        _normalized_element_id(raw)
        attrs: dict[str, Any] = {}
        for key in list(raw):
            if not key.startswith("attr_"):
                continue
            semantic = semantic_attribute_name(attribute_labels.get(key), key)
            if semantic in attrs:
                suffix = 2
                while f"{semantic}_{suffix}" in attrs:
                    suffix += 1
                semantic = f"{semantic}_{suffix}"
            attrs[semantic] = raw.pop(key)
        raw["attributes"] = attrs
        raw["source_row"] = row_index
        semantic_labels = {
            semantic_attribute_name(label, key): label for key, label in attribute_labels.items()
        }
        raw, row_findings = _security_scan(raw, field_labels=semantic_labels)
        findings.extend({**item, "sourceRow": row_index} for item in row_findings)
        records.append(raw)
    records.sort(key=stable_sort_key)
    populated = [record for record in records if _public_entity_populated(record)]
    missing = [record for record in records if not _public_entity_populated(record)]
    return {
        "records": records,
        "populated": populated,
        "missing": missing,
        "supplementalRows": supplemental_rows,
        "placeholderRows": placeholders,
        "unclassifiedRows": unclassified,
        "templateRows": sum(bool(row) for row in rows[:data_start]),
        "attributeLabels": attribute_labels,
        "errors": [],
        "securityFindings": findings,
    }


def _parse_metadata(rows: list[list[Any]]) -> dict[str, Any]:
    header_index, headers = _find_header(rows, {"element_id", "indicator_id"})
    if header_index is None:
        return {
            "records": [],
            "supplementalRows": [],
            "placeholderRows": 0,
            "unclassifiedRows": 0,
            "templateRows": sum(bool(row) for row in rows),
            "errors": ["metadata-header-not-found"],
            "securityFindings": [],
        }
    records: list[dict[str, Any]] = []
    placeholders = 0
    unclassified = 0
    supplemental_rows: list[dict[str, Any]] = []
    findings: list[dict[str, Any]] = []
    for row_index, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        if not row:
            continue
        candidate = _record_from_row(headers, row)
        if not _valid_source_record(candidate):
            if is_explicit_placeholder_row(row):
                placeholders += 1
            else:
                supplemental_rows.append({"source_row": row_index, "cells": row})
            continue
        _normalized_element_id(candidate)
        candidate["source_row"] = row_index
        candidate, row_findings = _security_scan(candidate)
        findings.extend({**item, "sourceRow": row_index} for item in row_findings)
        records.append(candidate)
    records.sort(key=stable_sort_key)
    return {
        "records": records,
        "supplementalRows": supplemental_rows,
        "placeholderRows": placeholders,
        "unclassifiedRows": unclassified,
        "templateRows": sum(bool(row) for row in rows[: header_index + 1]),
        "errors": [],
        "securityFindings": findings,
    }


def _parse_framework(rows: list[list[Any]]) -> list[dict[str, Any]]:
    header_index = None
    headers: list[str] = []
    for index, row in enumerate(rows[:25]):
        normalized = [normalize_field_name(value) for value in row]
        if "요소id" in normalized or "element_id" in normalized:
            header_index = index
            headers = ["element_id" if value == "요소id" else value for value in normalized]
            break
    if header_index is None:
        return []
    records: list[dict[str, Any]] = []
    for row in rows[header_index + 1 :]:
        if not row:
            continue
        record = _record_from_row(headers, row)
        element_id = extract_element_id(record.get("element_id"))
        if not element_id:
            continue
        record["element_id"] = element_id
        records.append(record)
    records.sort(key=lambda item: item["element_id"])
    return records


def _years_from_records(
    observations: Iterable[Mapping[str, Any]],
    metadata: Iterable[Mapping[str, Any]],
) -> list[int]:
    years: set[int] = set()
    for record in observations:
        value = record.get("year")
        if isinstance(value, int) and 1900 <= value <= 2100:
            years.add(value)
        elif value is not None:
            years.update(int(match) for match in _YEAR_RE.findall(str(value)))
    for record in metadata:
        for key in ("reference_year", "time_range", "last_updated", "accessed_date"):
            value = record.get(key)
            if value is not None:
                years.update(int(match) for match in _YEAR_RE.findall(str(value)))
    return sorted(year for year in years if 1900 <= year <= 2100)


def _unique_values(records: Iterable[Mapping[str, Any]], key: str) -> list[Any]:
    values: dict[str, Any] = {}
    for record in records:
        value = record.get(key)
        if is_placeholder(value):
            continue
        normalized = normalize_value(value)
        values[str(normalized)] = normalized
    return [values[key] for key in sorted(values)]


def parse_workbook_bytes(
    data: bytes,
    archive_name: str,
    *,
    include_records: bool = True,
) -> dict[str, Any]:
    """Parse one workbook, returning normalized records and audit metadata."""

    normalized_name = decode_zip_filename(archive_name)
    expected_element_id = extract_element_id(normalized_name)
    warnings: list[str] = []
    errors: list[str] = []
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:  # quarantine malformed workbooks without stopping the batch
        return {
            "archiveName": normalized_name,
            "elementId": expected_element_id,
            "sourceCategory": "E",
            "normalizationResult": "quarantined",
            "warnings": [],
            "errors": [f"workbook-load-error:{type(exc).__name__}:{exc}"],
        }
    try:
        missing_sheets = [sheet for sheet in CORE_SHEETS if sheet not in workbook.sheetnames]
        if missing_sheets:
            errors.extend(f"missing-sheet:{sheet}" for sheet in missing_sheets)

        sheet_data: dict[str, list[list[Any]]] = {}
        for sheet_name in CORE_SHEETS:
            if sheet_name not in workbook.sheetnames:
                sheet_data[sheet_name] = []
                continue
            rows, sheet_warnings = _sheet_rows(workbook[sheet_name])
            sheet_data[sheet_name] = rows
            warnings.extend(f"{sheet_name}:{warning}" for warning in sheet_warnings)

        observations = _parse_observations(sheet_data[OBSERVATION_SHEET])
        entities = _parse_entities(sheet_data[ENTITY_SHEET])
        metadata = _parse_metadata(sheet_data[METADATA_SHEET])
        framework = _parse_framework(sheet_data[FRAMEWORK_SHEET])
        errors.extend(observations["errors"] + entities["errors"] + metadata["errors"])

        all_records = observations["records"] + entities["records"] + metadata["records"]
        workbook_ids = sorted(
            {str(record.get("element_id")) for record in all_records if record.get("element_id")}
        )
        mismatched_ids = [item for item in workbook_ids if expected_element_id and item != expected_element_id]
        if mismatched_ids:
            errors.append("element-id-mismatch:" + ",".join(mismatched_ids))
        if not expected_element_id:
            errors.append("filename-element-id-missing")

        duplicate_observations = duplicate_summary(observations["records"])
        duplicate_entities = duplicate_summary(entities["records"])
        duplicate_metadata = duplicate_summary(metadata["records"])
        duplicate_count = sum(
            item["duplicateCount"]
            for item in (duplicate_observations, duplicate_entities, duplicate_metadata)
        )
        if duplicate_count:
            warnings.append(f"duplicate-records:{duplicate_count}")

        security_findings = (
            observations["securityFindings"]
            + entities["securityFindings"]
            + metadata["securityFindings"]
        )
        if security_findings:
            warnings.append(f"credential-values-removed:{len(security_findings)}")

        populated_count = len(observations["populated"]) + len(entities["populated"])
        placeholder_count = (
            observations["placeholderRows"]
            + entities["placeholderRows"]
            + metadata["placeholderRows"]
        )
        if errors:
            source_category = "E"
            normalization_result = "quarantined"
        elif populated_count:
            source_category = "A"
            normalization_result = "normalized"
        elif placeholder_count:
            source_category = "C"
            normalization_result = "explicit-placeholder-only"
        else:
            source_category = "B"
            normalization_result = "schema-only"

        years = _years_from_records(observations["records"], metadata["records"])
        spatial_units = _unique_values(metadata["records"], "spatial_unit")
        data_types = _unique_values(metadata["records"], "data_type")
        has_coordinates = any(
            not is_placeholder(record.get("lat")) and not is_placeholder(record.get("lon"))
            for record in entities["records"]
        )
        has_geometry = any(
            not is_placeholder(record.get("geometry_type")) for record in entities["records"]
        ) or any(str(value).startswith(("geo_", "raster")) for value in data_types)

        row_balance = {
            "observation": {
                "records": len(observations["records"]),
                "populated": len(observations["populated"]),
                "missing": len(observations["missing"]),
                "balanced": len(observations["records"])
                == len(observations["populated"]) + len(observations["missing"]),
            },
            "entity": {
                "records": len(entities["records"]),
                "populated": len(entities["populated"]),
                "missing": len(entities["missing"]),
                "balanced": len(entities["records"])
                == len(entities["populated"]) + len(entities["missing"]),
            },
        }
        row_balance["pass"] = row_balance["observation"]["balanced"] and row_balance["entity"]["balanced"]

        result: dict[str, Any] = {
            "archiveName": normalized_name,
            "elementId": expected_element_id,
            "elementIdsFound": workbook_ids,
            "sourceCategory": source_category,
            "normalizationResult": normalization_result,
            "sheetNames": [nfc_text(name) for name in workbook.sheetnames],
            "observationRowCount": len(observations["records"]),
            "observationPopulatedRowCount": len(observations["populated"]),
            "observationMissingRowCount": len(observations["missing"]),
            "entityRowCount": len(entities["records"]),
            "entityPopulatedRowCount": len(entities["populated"]),
            "entityMissingRowCount": len(entities["missing"]),
            "metadataRowCount": len(metadata["records"]),
            "templateRowCount": (
                observations["templateRows"]
                + entities["templateRows"]
                + metadata["templateRows"]
            ),
            "placeholderRowCount": placeholder_count,
            "unclassifiedNonemptyRowCount": (
                observations["unclassifiedRows"]
                + entities["unclassifiedRows"]
                + metadata["unclassifiedRows"]
            ),
            "supplementalSourceRowCount": (
                len(observations["supplementalRows"])
                + len(entities["supplementalRows"])
                + len(metadata["supplementalRows"])
            ),
            "publicPopulatedRowCount": populated_count,
            "latestYear": max(years) if years else None,
            "years": years,
            "sourceOrganizations": _unique_values(metadata["records"], "source_org"),
            "sourceUrls": _unique_values(metadata["records"], "source_url"),
            "originalLicenses": _unique_values(metadata["records"], "license_code"),
            "attributionTexts": _unique_values(metadata["records"], "attribution_text"),
            "dataTypes": data_types,
            "spatialUnits": spatial_units,
            "coordinateOrGeometry": has_coordinates or has_geometry,
            "hasCoordinates": has_coordinates,
            "hasGeometry": has_geometry,
            "duplicateDetection": {
                "observation": duplicate_observations,
                "entity": duplicate_entities,
                "metadata": duplicate_metadata,
                "duplicateCount": duplicate_count,
            },
            "credentialDetection": {
                "removedCount": len(security_findings),
                "findings": security_findings,
                "normalContactsSuppressed": 0,
            },
            "excelErrorCellCount": sum("excel-error-cell:" in warning for warning in warnings),
            "rowBalance": row_balance,
            "warnings": sorted(set(warnings)),
            "errors": sorted(set(errors)),
        }
        if include_records:
            result["observations"] = observations["records"]
            result["entities"] = entities["records"]
            result["metadata"] = metadata["records"]
            result["supplementalSourceRows"] = {
                "observation": observations["supplementalRows"],
                "entity": entities["supplementalRows"],
                "metadata": metadata["supplementalRows"],
            }
            result["framework"] = framework
            result["entityAttributeLabels"] = entities["attributeLabels"]
        else:
            result["frameworkElementCount"] = len(framework)
        return result
    finally:
        workbook.close()
